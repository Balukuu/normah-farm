"""Builds crop-specifications.xlsx from data/crops.json — the single
source of truth also used by the website's crop pages and comparison
table. Fields the client must confirm are left blank and highlighted
amber, never guessed. Re-run this after editing data/crops.json to
keep the spreadsheet in sync."""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_PATH = os.path.join(BASE, "data", "crops.json")
OUT_PATH = os.path.join(BASE, "assets", "docs", "crop-specifications.xlsx")

with open(DATA_PATH, encoding="utf-8") as f:
    data = json.load(f)

SEED = "1E6B3A"
CHAFF_FILL = "FCE9A8"  # soft tint of --chaff, for "needs client input" cells

wb = Workbook()
ws = wb.active
ws.title = "Crop specifications"

headers = [
    "Crop", "Botanical name", "Status", "Season(s)", "Season note",
    "Planting window", "Harvest window", "Moisture target (%)",
    "Variety", "Packaging", "Unit", "Minimum order",
    "Primary market", "Notes",
]
ws.append(headers)

header_fill = PatternFill("solid", fgColor=SEED)
header_font = Font(color="FBFCFA", bold=True, name="Calibri", size=11)
thin = Side(style="thin", color="D9DED6")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, _ in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 30

VERIFY_FILL = PatternFill("solid", fgColor=CHAFF_FILL)

season_map = data["seasons"]

def season_label(codes):
    if not codes:
        return "Planned — not yet assigned"
    return " & ".join(season_map[c]["label"].replace("Season ", "") for c in codes)

row_i = 2
for crop in data["crops"]:
    seasons_txt = season_label(crop.get("seasons", []))
    season_note = crop.get("seasonNote", {}).get("value", "") or ""
    season_note_verify = crop.get("seasonNote", {}).get("verify", False)

    def val(key):
        field = crop.get(key)
        if field is None:
            return "", False
        return field.get("value") or "", field.get("verify", False)

    botanical, botanical_v = val("botanicalName")
    planting, planting_v = val("plantingWindow")
    harvest, harvest_v = val("harvestWindow")
    moisture, moisture_v = val("moistureTarget")
    variety, variety_v = val("variety")
    packaging, packaging_v = val("packaging")
    unit, unit_v = val("unit")
    min_order, min_order_v = val("minimumOrder")

    row = [
        crop["name"],
        botanical,
        crop["status"],
        seasons_txt,
        season_note,
        planting,
        harvest,
        moisture,
        variety,
        packaging,
        unit,
        min_order,
        crop.get("primaryMarket", ""),
        crop.get("bodyNote", crop.get("useNote", "")),
    ]
    ws.append(row)

    verify_flags = [
        False, botanical_v, False, False, season_note_verify,
        planting_v, harvest_v, moisture_v, variety_v, packaging_v,
        unit_v, min_order_v, False, False,
    ]
    for col_idx, flag in enumerate(verify_flags, start=1):
        cell = ws.cell(row=row_i, column=col_idx)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if flag:
            cell.fill = VERIFY_FILL
    row_i += 1

widths = [14, 16, 10, 12, 26, 20, 20, 16, 16, 22, 12, 14, 26, 40]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

legend = wb.create_sheet("Read me")
legend["A1"] = "Normah Agro Farm — crop specifications"
legend["A1"].font = Font(bold=True, size=14, color=SEED)
legend["A3"] = (
    "This file is the single source of truth for the crop specification "
    "content on normahagrofarm.com (see /data/crops.json) and for the "
    "downloadable product specification PDF."
)
legend["A3"].alignment = Alignment(wrap_text=True)
legend.merge_cells("A3:F3")
legend.row_dimensions[3].height = 40

legend["A5"] = "Amber cells"
legend["A5"].font = Font(bold=True)
legend["B5"] = "Need the client to confirm a value. Nothing in an amber cell is a guess — fill it in, then update /data/crops.json to match so the website and PDF stay in sync."
legend["B5"].alignment = Alignment(wrap_text=True)
legend.merge_cells("B5:F5")
legend.row_dimensions[5].height = 45
legend["A5"].fill = VERIFY_FILL

legend.column_dimensions["A"].width = 16
legend.column_dimensions["B"].width = 70

wb.save(OUT_PATH)
print("wrote", OUT_PATH)
